from __future__ import annotations

import re
import unicodedata
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from apps.catalogo.models import Categoria, Marca, Presentacion, Producto, Proveedor, UnidadMedida
from apps.precios.models import ListaPrecio, PrecioProducto

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise CommandError(
        'Falta openpyxl. Instala dependencias con: pip install -r requirements.txt'
    ) from exc


BATCH_SIZE = 500


def normalize_header(value: Any) -> str:
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r'[^a-z0-9]+', '', text)


def cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal) and value == value.to_integral_value():
        return str(int(value))
    return str(value).strip()


def parse_money(value: Any) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = str(value).strip()
    if not raw:
        return None
    if re.search(r',\d{1,2}$', raw) and '.' in raw:
        raw = raw.replace('.', '').replace(',', '.')
    else:
        raw = raw.replace(',', '')
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def parse_bool_si_no(value: Any, default: bool = True) -> bool:
    text = cell_str(value)
    if not text:
        return default
    return text.upper() in {'SI', 'S', 'YES', 'TRUE', '1'}


# Listas internas de precios: no traen las columnas ocultas de SIIGO (categoría, tipo, etc.).
SHEETS_PRECIO_SIN_SIIGO = {'MEDICAMENTOS', 'PAPELERIA'}

# Orden de cada set = prioridad (el alias más específico gana, aunque esté más a la derecha).
HEADER_ALIASES = {
    'sku': {
        'codigodelproductoobligatorio',
        'codigodelproducto',
        'codigoproducto',
        'codigo',
    },
    'nombre': {
        'nombredelproductoservicioobligatorio',
        'nombredelproductoservicio',
        'nombredelproducto',
        'producto',
        'nombre',
    },
    'categoria': {
        'categoriadeinventariosserviciosobligatorio',
        'categoriadeinventariosservicios',
        'categoriadeinventario',
        'categoria',
    },
    'tipo': {
        'tipodeproductoobligatorio',
        'tipodeproducto',
    },
    'inventariable': {
        'inventariableobligatorio',
        'inventariable',
    },
    'referencia': {
        'referenciadefabrica',
        'referencia',
    },
    'descripcion': {
        'descripcionlarga',
        'descripcion',
    },
    'unidad_impresion': {
        'unidaddemedidaimpresionfactura',
        'unidadimpresion',
    },
    'precio': {
        'valorventaconsumidorfinal',
        'totalconsumidorfinalaldetal',
        'valorconiva',
        'valortotal',
        'valorunitario',
        'precio',
        'valor',
    },
}

# Si el alias exacto falla (tildes/ocultas), exige que el header contenga estas piezas.
HEADER_CONTAINS = {
    'categoria': ('categoria',),
    'tipo': ('tipodeproducto',),
    'inventariable': ('inventariable',),
}


def detect_column_map(header_row: tuple[Any, ...]) -> dict[str, int] | None:
    """Mapea encabezados visibles u ocultos. Prioriza alias específicos, no el índice de columna."""
    normalized = [normalize_header(cell) for cell in header_row]
    column_map: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            for idx, header in enumerate(normalized):
                if header == alias:
                    column_map[field] = idx
                    break
            if field in column_map:
                break
        if field in column_map:
            continue
        needles = HEADER_CONTAINS.get(field)
        if not needles:
            continue
        for idx, header in enumerate(normalized):
            if header and all(needle in header for needle in needles):
                column_map[field] = idx
                break
    if 'sku' in column_map and 'nombre' in column_map:
        return column_map
    return None


def proveedor_from_sheet_name(sheet_name: str) -> str:
    cleaned = sheet_name.strip()
    match = re.match(r'^\d+\s*-?\s*(.+)$', cleaned)
    if match:
        return match.group(1).strip()
    return cleaned


class Command(BaseCommand):
    help = (
        'Importa productos desde un Excel SIIGO con múltiples hojas. '
        'Detecta automáticamente el formato de cada hoja.'
    )

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Ruta al archivo .xlsx')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Valida e informa sin guardar cambios',
        )
        parser.add_argument(
            '--sheet',
            action='append',
            dest='sheets',
            help='Procesar solo hojas específicas (se puede repetir). Por defecto: todas.',
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options['xlsx_path'])
        if not xlsx_path.exists():
            raise CommandError(f'No existe el archivo: {xlsx_path}')
        if xlsx_path.suffix.lower() not in {'.xlsx', '.xlsm'}:
            raise CommandError('El archivo debe ser Excel (.xlsx)')

        dry_run = options['dry_run']
        only_sheets = set(options['sheets'] or [])

        self.stdout.write(f'Abriendo {xlsx_path.name}...')
        # Primero solo leemos nombres (archivo grande: no mantener el libro abierto).
        bootstrap = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_names = list(bootstrap.sheetnames)
        bootstrap.close()

        totals = {
            'sheets_ok': 0,
            'sheets_skip': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }

        try:
            if dry_run:
                with transaction.atomic():
                    totals = self._process_sheets(
                        xlsx_path=xlsx_path,
                        sheet_names=sheet_names,
                        only_sheets=only_sheets,
                        totals=totals,
                    )
                    raise CommandError('Dry-run: rollback intencional')
            totals = self._process_sheets(
                xlsx_path=xlsx_path,
                sheet_names=sheet_names,
                only_sheets=only_sheets,
                totals=totals,
            )
        except CommandError as exc:
            if dry_run and 'Dry-run' in str(exc):
                self.stdout.write(self.style.WARNING('Dry-run completado (sin guardar).'))
            else:
                raise

        self.stdout.write(
            self.style.SUCCESS(
                'Importación finalizada — '
                f"hojas OK: {totals['sheets_ok']}, hojas omitidas: {totals['sheets_skip']}, "
                f"creados: {totals['created']}, actualizados: {totals['updated']}, "
                f"omitidos: {totals['skipped']}, errores: {totals['errors']}"
            )
        )

    def _process_sheets(
        self,
        *,
        xlsx_path: Path,
        sheet_names: list[str],
        only_sheets: set[str],
        totals: dict[str, int],
    ) -> dict[str, int]:
        context = self._build_context()
        total_sheets = len(sheet_names)
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            for sheet_index, sheet_name in enumerate(sheet_names, start=1):
                if only_sheets and sheet_name not in only_sheets:
                    continue

                if not only_sheets and sheet_name.strip().upper() in SHEETS_PRECIO_SIN_SIIGO:
                    totals['sheets_skip'] += 1
                    self.stdout.write(
                        f'[{sheet_index}/{total_sheets}] {sheet_name} — '
                        'omitida (lista de precios; las columnas SIIGO ocultas están en las hojas de proveedor)'
                    )
                    continue

                self.stdout.write(f'[{sheet_index}/{total_sheets}] {sheet_name}')
                worksheet = workbook[sheet_name]
                with transaction.atomic():
                    result = self._import_sheet(
                        sheet_name,
                        worksheet,
                        context,
                        require_categoria=not bool(only_sheets),
                    )

                if result['usable']:
                    totals['sheets_ok'] += 1
                    self.stdout.write(
                        self.style.SUCCESS(
                            f"  ✓ +{result['created']} ~{result['updated']} "
                            f"omitidos={result['skipped']} errores={result['errors']}"
                        )
                    )
                else:
                    totals['sheets_skip'] += 1

                totals['created'] += result['created']
                totals['updated'] += result['updated']
                totals['skipped'] += result['skipped']
                totals['errors'] += result['errors']
        finally:
            workbook.close()

        return totals

    def _build_context(self) -> dict[str, Any]:
        lista_precio, _ = ListaPrecio.objects.get_or_create(
            nombre='General',
            defaults={
                'descripcion': 'Lista de precios por defecto (importación SIIGO)',
                'es_default': True,
                'activo': True,
            },
        )
        if not lista_precio.es_default:
            ListaPrecio.objects.filter(es_default=True).exclude(pk=lista_precio.pk).update(
                es_default=False
            )
            lista_precio.es_default = True
            lista_precio.save(update_fields=['es_default'])

        unidad, _ = UnidadMedida.objects.get_or_create(
            abreviatura='UND',
            defaults={'nombre': 'Unidad', 'activo': True},
        )

        productos_by_sku = {
            sku: pk for sku, pk in Producto.all_objects.values_list('sku', 'id')
        }

        return {
            'lista_precio': lista_precio,
            'unidad': unidad,
            'categoria_cache': {},
            'marca_cache': {},
            'proveedor_cache': {},
            'presentacion_cache': {},
            'productos_by_sku': productos_by_sku,
        }

    def _import_sheet(
        self,
        sheet_name: str,
        worksheet,
        context: dict[str, Any],
        *,
        require_categoria: bool = True,
    ) -> dict[str, Any]:
        stats = {
            'usable': False,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }

        column_map = None
        sheet_proveedor = proveedor_from_sheet_name(sheet_name)
        pending_create: list[Producto] = []
        pending_prices: list[tuple[str, Decimal]] = []  # sku, precio
        pending_updates: list[Producto] = []
        processed = 0

        def flush():
            nonlocal pending_create, pending_updates, pending_prices
            if pending_create:
                Producto.objects.bulk_create(pending_create, batch_size=BATCH_SIZE)
                created_skus = [product.sku for product in pending_create]
                for sku, pk in Producto.all_objects.filter(sku__in=created_skus).values_list('sku', 'id'):
                    context['productos_by_sku'][sku] = pk
                stats['created'] += len(pending_create)
                pending_create = []

            if pending_updates:
                Producto.objects.bulk_update(
                    pending_updates,
                    fields=[
                        'nombre',
                        'descripcion_corta',
                        'descripcion_larga',
                        'categoria',
                        'marca',
                        'proveedor',
                        'presentacion',
                        'referencia_fabricante',
                        'tipo_producto',
                        'inventariable',
                        'activo',
                        'deleted_at',
                    ],
                    batch_size=BATCH_SIZE,
                )
                stats['updated'] += len(pending_updates)
                pending_updates = []

            if pending_prices:
                self._upsert_prices(pending_prices, context)
                pending_prices = []

        for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if column_map is None:
                detected = detect_column_map(row)
                if detected:
                    if 'categoria' not in detected and require_categoria:
                        self.stdout.write(
                            '  omitida (sin columna Categoría de Inventarios, suele estar oculta)'
                        )
                        break
                    column_map = detected
                    stats['usable'] = True
                    self.stdout.write(
                        '  columnas: ' + ', '.join(f'{k}={v + 1}' for k, v in column_map.items())
                    )
                    continue
                if idx >= 15:
                    break
                continue

            if not row or all(cell is None or cell_str(cell) == '' for cell in row):
                stats['skipped'] += 1
                continue

            try:
                parsed = self._parse_row(row, column_map, sheet_proveedor, context)
            except Exception as exc:  # noqa: BLE001
                stats['errors'] += 1
                self.stderr.write(f'  ! {sheet_name} fila {idx}: {exc}')
                continue

            if parsed is None:
                stats['skipped'] += 1
                continue

            sku = parsed.pop('sku')
            precio = parsed.pop('precio')

            if sku in context['productos_by_sku']:
                existing_id = context['productos_by_sku'][sku]
                if existing_id is None:
                    # Ya va en el batch de creación: solo refrescar precio
                    if precio is not None:
                        pending_prices.append((sku, precio))
                    stats['skipped'] += 1
                    continue
                product = Producto(id=existing_id, sku=sku, **parsed)
                pending_updates.append(product)
            else:
                product = Producto(sku=sku, **parsed)
                pending_create.append(product)
                context['productos_by_sku'][sku] = None

            if precio is not None:
                pending_prices.append((sku, precio))

            processed += 1
            if len(pending_create) + len(pending_updates) >= BATCH_SIZE:
                flush()
                self.stdout.write(f'  ... {sheet_name}: {processed} filas procesadas')

        flush()

        if stats['usable'] and stats['created'] == 0 and stats['updated'] == 0 and stats['errors'] == 0:
            stats['usable'] = False
        return stats

    def _upsert_prices(self, pending_prices: list[tuple[str, Decimal]], context: dict[str, Any]) -> None:
        lista = context['lista_precio']
        today = date.today()
        skus = [sku for sku, _ in pending_prices]
        products = {
            p.sku: p
            for p in Producto.all_objects.filter(sku__in=skus).only('id', 'sku')
        }

        existing = {
            (precio.producto_id, precio.lista_precio_id): precio
            for precio in PrecioProducto.objects.filter(
                producto_id__in=[p.id for p in products.values()],
                lista_precio=lista,
                fecha_inicio=today,
            )
        }

        to_create: list[PrecioProducto] = []
        to_update: list[PrecioProducto] = []

        for sku, amount in pending_prices:
            product = products.get(sku)
            if product is None:
                continue
            key = (product.id, lista.id)
            current = existing.get(key)
            if current is None:
                obj = PrecioProducto(
                    producto=product,
                    lista_precio=lista,
                    precio_base=amount,
                    moneda='COP',
                    fecha_inicio=today,
                    activo=True,
                )
                to_create.append(obj)
                existing[key] = obj
            else:
                current.precio_base = amount
                current.activo = True
                to_update.append(current)

        if to_create:
            PrecioProducto.objects.bulk_create(to_create, batch_size=BATCH_SIZE)
        if to_update:
            PrecioProducto.objects.bulk_update(
                to_update, fields=['precio_base', 'activo'], batch_size=BATCH_SIZE
            )

    def _parse_row(
        self,
        row: tuple[Any, ...],
        column_map: dict[str, int],
        sheet_proveedor: str,
        context: dict[str, Any],
    ) -> dict[str, Any] | None:
        def col(field: str) -> Any:
            idx = column_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        sku = cell_str(col('sku'))[:50]
        nombre = cell_str(col('nombre'))
        if not sku or not nombre:
            return None
        if normalize_header(sku) in HEADER_ALIASES['sku'] and normalize_header(nombre) in HEADER_ALIASES['nombre']:
            return None

        categoria_raw = cell_str(col('categoria'))
        categoria_nombre = self._nombre_categoria(categoria_raw)

        referencia_fabricante = cell_str(col('referencia')) or None
        proveedor_nombre = referencia_fabricante or sheet_proveedor or 'Proveedor SIIGO'

        return {
            'sku': sku,
            'nombre': nombre[:255],
            'descripcion_corta': nombre[:255],
            'descripcion_larga': cell_str(col('descripcion')) or None,
            'categoria': self._get_categoria(categoria_nombre, context['categoria_cache']),
            'marca': self._get_marca(self._guess_marca(proveedor_nombre), context['marca_cache']),
            'proveedor': self._get_proveedor(proveedor_nombre, context['proveedor_cache']),
            'presentacion': self._get_presentacion(
                cell_str(col('unidad_impresion')),
                context['unidad'],
                context['presentacion_cache'],
            ),
            'referencia_fabricante': (referencia_fabricante or '')[:100] or None,
            'tipo_producto': self._map_tipo_producto(cell_str(col('tipo')), categoria_nombre),
            'inventariable': parse_bool_si_no(col('inventariable'), default=True),
            'activo': True,
            'deleted_at': None,
            'precio': parse_money(col('precio')),
        }

    @staticmethod
    def _nombre_categoria(raw: str) -> str:
        """'3-Medicamentos' / '5 Cuidado Personal' → nombre limpio."""
        text = (raw or '').strip()
        if not text:
            return 'General'
        match = re.match(r'^\d+\s*[-–:]?\s*(.+)$', text)
        nombre = (match.group(1) if match else text).strip(' -–:')
        if not nombre or re.fullmatch(r'\d+', nombre):
            return 'General'
        return nombre

    def _get_categoria(self, nombre: str, cache: dict[str, Categoria]) -> Categoria:
        key = nombre.casefold()
        if key not in cache:
            slug = slugify(nombre)[:150] or 'general'
            cache[key], _ = Categoria.objects.get_or_create(
                slug=slug,
                defaults={'nombre': nombre[:150], 'nivel': 1, 'activo': True},
            )
        return cache[key]

    def _get_marca(self, nombre: str, cache: dict[str, Marca]) -> Marca:
        key = nombre.casefold()
        if key not in cache:
            cache[key], _ = Marca.objects.get_or_create(
                nombre=nombre[:150],
                defaults={'activo': True},
            )
        return cache[key]

    def _get_proveedor(self, nombre: str, cache: dict[str, Proveedor]) -> Proveedor:
        key = nombre.casefold()
        if key not in cache:
            nit_key = slugify(nombre).replace('-', '')[:20] or 'SIIGO'
            cache[key], _ = Proveedor.objects.get_or_create(
                nit=nit_key,
                defaults={'nombre': nombre[:200], 'activo': True},
            )
        return cache[key]

    def _get_presentacion(
        self,
        unidad_impresion: str,
        unidad: UnidadMedida,
        cache: dict[str, Presentacion],
    ) -> Presentacion | None:
        if not unidad_impresion:
            return None
        if unidad_impresion not in cache:
            cache[unidad_impresion], _ = Presentacion.objects.get_or_create(
                nombre=f'Unidad impresión {unidad_impresion}'[:100],
                defaults={
                    'unidad_medida': unidad,
                    'cantidad': Decimal('1'),
                    'activo': True,
                },
            )
        return cache[unidad_impresion]

    @staticmethod
    def _guess_marca(proveedor_nombre: str) -> str:
        name = proveedor_nombre.strip()
        upper = name.upper()
        known = {
            'ALCON': 'Alcon',
            'BAYER': 'Bayer',
            'ABBOTT': 'Abbott',
            'PFIZER': 'Pfizer',
            'NIVEA': 'Nivea',
            '3M': '3M',
        }
        for token, brand in known.items():
            if token in upper:
                return brand
        for suffix in (' S. A. S.', ' S.A.S.', ' S. A.', ' S.A.', ' SAS', ' LTDA', ' LTD'):
            if name.upper().endswith(suffix.upper()):
                name = name[: -len(suffix)].strip()
                break
        return name[:150] or 'Sin marca'

    @staticmethod
    def _map_tipo_producto(tipo_raw: str, categoria_nombre: str) -> str:
        cat = categoria_nombre.upper()
        if 'MEDICAMENT' in cat:
            return Producto.TipoProducto.MEDICAMENTO
        if 'COSMET' in cat:
            return Producto.TipoProducto.COSMETICO
        if 'DISPOSIT' in cat or 'EQUIPO' in cat:
            return Producto.TipoProducto.DISPOSITIVO
        if tipo_raw.upper().startswith('P-'):
            return Producto.TipoProducto.OTRO
        return Producto.TipoProducto.OTRO
